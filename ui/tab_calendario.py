import datetime
import streamlit as st
from utils.content_utils import construir_contexto

INFO_SAFRAS = """
Calendário de Safra: Algodão
Tocantins: Plantio de novembro (2ª quinzena) até fevereiro (2ª quinzena), com pico intenso em janeiro. Colheita de abril (2ª quinzena) até agosto (1ª quinzena), com pico intenso em junho e julho.
Maranhão: Plantio de dezembro (1ª quinzena) até março (2ª quinzena), com pico intenso em janeiro. Colheita de maio (2ª quinzena) até agosto (2ª quinzena), com pico intenso em junho e julho.
Mato Grosso: Plantio de dezembro (1ª quinzena) até fevereiro (2ª quinzena), com pico intenso em janeiro. Colheita de abril (2ª quinzena) até agosto (2ª quinzena), com pico intenso em junho.
São Paulo: Plantio de outubro (1ª quinzena) até janeiro (1ª quinzena), com pico intenso em novembro e dezembro. Colheita de março (1ª quinzena) até junho (1ª quinzena), com pico intenso em abril e maio.
Paraná: Plantio de setembro (2ª quinzena) até dezembro (1ª quinzena), com pico intenso em outubro e novembro. Colheita de março (1ª quinzena) até maio (2ª quinzena), com pico intenso em abril.

Calendário de Safra: Soja
Mato Grosso: Plantio de setembro (1ª quinzena) até dezembro (2ª quinzena), com pico intenso em outubro e novembro. Colheita de janeiro (1ª quinzena) até abril (1ª quinzena), com pico intenso em fevereiro e março.
Mato Grosso do Sul: Plantio de setembro (2ª quinzena) até dezembro (2ª quinzena), com pico intenso em outubro e novembro. Colheita de janeiro (2ª quinzena) até abril (1ª quinzena), com pico intenso em março.
Goiás: Plantio de setembro (2ª quinzena) até janeiro (1ª quinzena), com pico intenso em outubro e novembro. Colheita de janeiro (2ª quinzena) até abril (1ª quinzena), com pico intenso em março.
Paraná: Plantio de setembro (2ª quinzena) até dezembro (2ª quinzena), com pico intenso em outubro e novembro. Colheita de janeiro (2ª quinzena) até abril (2ª quinzena), com pico intenso em março.
Rio Grande do Sul: Plantio de outubro (1ª quinzena) até janeiro (1ª quinzena), com pico intenso em novembro e dezembro. Colheita de fevereiro (1ª quinzena) até maio (1ª quinzena), com pico intenso em março e abril.

Calendário de Safra: Milho 1ª Safra
Mato Grosso: Plantio de setembro (2ª quinzena) até dezembro (1ª quinzena), com pico intenso em outubro e novembro. Colheita de fevereiro (1ª quinzena) até maio (1ª quinzena), com pico intenso em março e abril.
Paraná: Plantio de agosto (2ª quinzena) até dezembro (1ª quinzena), com pico intenso em setembro e outubro. Colheita de janeiro (2ª quinzena) até junho (2ª quinzena), com pico intenso em março.

Calendário de Safra: Milho 2ª Safra
Mato Grosso: Plantio de janeiro (2ª quinzena) até março (1ª quinzena), com pico intenso em fevereiro. Colheita de maio (2ª quinzena) até agosto (1ª quinzena), com pico intenso em junho e julho.
Goiás: Plantio de janeiro (1ª quinzena) até março (1ª quinzena), com pico intenso em fevereiro. Colheita de maio (1ª quinzena) até setembro (1ª quinzena), com pico intenso em junho e julho.

Calendário de Safra: Cana-de-Açúcar
Centro-Oeste: Plantio de janeiro (1ª quinzena) até julho (1ª quinzena) e de outubro (1ª quinzena) até dezembro (2ª quinzena). Colheita de abril (1ª quinzena) até novembro (2ª quinzena).
Sudeste: Plantio de janeiro (1ª quinzena) até julho (1ª quinzena) e de outubro (1ª quinzena) até dezembro (2ª quinzena). Colheita de abril (1ª quinzena) até novembro (2ª quinzena).
"""


def render(tab, modelo_texto):
    with tab:
        st.header("📅 Criadora de Calendário")

        if not st.session_state.agente_selecionado:
            st.warning("Nenhum agente selecionado.")
            return

        agente = st.session_state.agente_selecionado
        st.success(f"Agente: {agente['nome']}")

        col1, col2 = st.columns([2, 1])

        with col1:
            mes_ano = st.text_input("Mês/Ano:", "FEVEREIRO 2026")
            data_inicio = st.date_input("Data início:", value=datetime.date(2026, 2, 1))
            data_fim = st.date_input("Data fim:", value=datetime.date(2026, 2, 28))
            delta_dias = (data_fim - data_inicio).days + 1

            culturas_prioritarias = st.text_area(
                "Culturas (separadas por vírgula, use 'e' para múltiplas):",
                "Soja, Milho, Cana-de-açúcar, Algodão, Soja e Milho, Soja e Cana",
            )
            culturas_lista = [c.strip() for c in culturas_prioritarias.split(",") if c.strip()]

        with col2:
            dias_com_1_pauta = st.number_input("Dias com 1 pauta:", 0, delta_dias, 5)
            dias_com_2_pautas = st.number_input("Dias com 2 pautas:", 0, delta_dias, 15)
            dias_com_3_pautas = st.number_input("Dias com 3 pautas:", 0, delta_dias, 3)
            dias_sem_pautas = delta_dias - (dias_com_1_pauta + dias_com_2_pautas + dias_com_3_pautas)
            if dias_sem_pautas < 0:
                st.error("Total excede dias disponíveis")

        st.subheader("Produtos e Direcionais")
        st.write("Formato: Produto(s) - Cultura(s) - Tema")

        produtos_direcionais = st.text_area(
            "Produtos com culturas e temas:",
            """Verdavis, Megafol e Victrato - Soja e Milho - Tecnologia para feira
Elestal Neo - Soja - Controle de mosca-branca
Fortenza - Milho - Seedcare para cigarrinha
YieldOn - Soja - Bioativador para pegamento
Miravis - Soja - Fungicida para ferrugem
Victrato - Cana - Nematicida para cana-soca
Victrato pelo Brasil - Soja e Cana - Ação nacional""",
            height=150,
        )

        produtos_com_direcionais = []
        if produtos_direcionais:
            for linha in produtos_direcionais.split('\n'):
                linha = linha.strip()
                if linha and ' - ' in linha:
                    partes = linha.split(' - ')
                    if len(partes) >= 3:
                        produtos_com_direcionais.append({
                            'produtos': [p.strip() for p in partes[0].split(' e ') if p.strip()],
                            'culturas': [c.strip() for c in partes[1].split(' e ') if c.strip()],
                            'tema': ' - '.join(partes[2:]).strip(),
                        })

        col_feira, col_recorrente = st.columns(2)

        with col_feira:
            st.write("Semana com evento (1 post/dia):")
            semana_feira_inicio = st.date_input("Início:", value=datetime.date(2026, 2, 9))
            semana_feira_fim = st.date_input("Fim:", value=datetime.date(2026, 2, 13))
            produtos_prioritarios_feira = st.text_input("Produtos prioritários:", "Verdavis, Megafol, Victrato")

        with col_recorrente:
            pauta_recorrente_texto = st.text_input("Pauta fixa:", "Victrato pelo Brasil")
            pauta_recorrente_dias = st.multiselect("Dias da semana:", ["Terça", "Quinta"], default=["Terça", "Quinta"])

        contexto_mensal = st.text_area(
            "Contexto do mês:",
            """FEVEREIRO 2026:
- Soja: colheita no centro-sul
- Milho: plantio da safrinha
- Cana: crescimento vegetativo
- Evento: Feira Nacional do Agronegócio (09-13/02)
- Foco: Verdavis, Megafol, Victrato na feira
- Pauta fixa: Victrato pelo Brasil (terças e quintas)""",
            height=120,
        )

        evitar_consecutivos_sem_pautas = st.checkbox("Evitar dias consecutivos sem pautas", True)
        max_repeticoes_tema = st.slider("Máx repetições por tema:", 1, 5, 2)

        if st.button("Gerar Calendário", type="primary"):
            if data_inicio >= data_fim:
                st.error("Data início deve ser anterior")
                return
            if not culturas_lista:
                st.error("Digite culturas")
                return
            if (dias_com_1_pauta + dias_com_2_pautas + dias_com_3_pautas) > delta_dias:
                st.error("Total excede período")
                return

            with st.spinner("Gerando calendário..."):
                try:
                    contexto_agente = construir_contexto(agente, st.session_state.segmentos_selecionados)

                    info_especifica = f"""
                    CONFIGURAÇÕES:
                    1. SEMANA COM EVENTO ({semana_feira_inicio.strftime('%d/%m')} a {semana_feira_fim.strftime('%d/%m')}):
                       - Apenas 1 pauta por dia
                       - Priorizar: {produtos_prioritarios_feira}

                    2. PAUTA FIXA: "{pauta_recorrente_texto}"
                       - Dias: {', '.join(pauta_recorrente_dias)}

                    3. FREQUÊNCIA:
                       - Dias com 1 pauta: {dias_com_1_pauta}
                       - Dias com 2 pautas: {dias_com_2_pautas}
                       - Dias com 3 pautas: {dias_com_3_pautas}
                       - Dias sem pautas: {max(0, dias_sem_pautas)}
                       - Evitar consecutivos sem pautas: {evitar_consecutivos_sem_pautas}

                    4. CONTROLE REPETIÇÃO:
                       - Máximo repetições por tema: {max_repeticoes_tema}
                    """

                    prompt_calendario = f'''
                    {contexto_agente}

                    ### BEGIN DADOS_SAFRA ###
                    {INFO_SAFRAS}
                    ### END DADOS_SAFRA ###

                    GERAR CALENDÁRIO COM ESTAS REGRAS:

                    PERÍODO: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}
                    MÊS: {mes_ano}

                    {info_especifica}

                    CONTEXTO: {contexto_mensal}

                    PRODUTOS E TEMAS:
                    {chr(10).join([f"- {', '.join(p['produtos'])} - {', '.join(p['culturas'])} - {p['tema']}" for p in produtos_com_direcionais])}

                    REGRAS CRÍTICAS:
                    1. Semana {semana_feira_inicio.strftime('%d/%m')} a {semana_feira_fim.strftime('%d/%m')}: APENAS 1 PAUTA POR DIA
                    2. Priorizar produtos: {produtos_prioritarios_feira} na semana da feira
                    3. Inserir "{pauta_recorrente_texto}" em TODAS as {', '.join(pauta_recorrente_dias)}
                    4. NÃO repetir temas (máximo {max_repeticoes_tema} repetições)
                    5. Células podem ter múltiplas culturas: "Soja e Milho", "Verdavis e Megafol"
                    6. Praticamente todos os dias com conteúdo
                    7. NUNCA 3 dias consecutivos sem pautas
                    8. Baseie pautas no contexto do mês
                    9. As pautas devem respeitar COM RIGIDEZ as fases reais de cada cultura por estados descritos no bloco 'DADOS_SAFRA'

                    FORMATO:
                    - Célula: "[EMOJI] Produto(s) - Cultura(s) - Tema - Breve descrição"
                    - Ex: "🔵 Verdavis e Megafol - Soja e Milho - Tecnologia feira - Soluções apresentadas na feira"
                    - Ex: "🟢 Victrato pelo Brasil - Soja e Cana - Ação nacional - Resultados em diferentes regiões"

                    Retorne CSV pronto para Excel.
                    '''

                    resposta = modelo_texto.generate_content(prompt_calendario)
                    calendario_csv = resposta.text.strip()

                    for tag in ['```csv', '```']:
                        calendario_csv = calendario_csv.replace(tag, '')

                    st.session_state.calendario_gerado = calendario_csv
                    st.session_state.mes_ano_calendario = mes_ano
                    st.success("Calendário gerado")

                except Exception as e:
                    st.error(f"Erro: {str(e)}")

        if 'calendario_gerado' in st.session_state:
            st.subheader(f"Calendário - {st.session_state.mes_ano_calendario}")

            tab_csv, tab_xlsx = st.tabs(["CSV", "XLSX"])

            with tab_csv:
                st.text_area("CSV:", st.session_state.calendario_gerado, height=400)
                st.download_button(
                    "Baixar CSV",
                    data=st.session_state.calendario_gerado,
                    file_name=f"calendario_{st.session_state.mes_ano_calendario.replace(' ', '_').lower()}.csv",
                    mime="text/csv",
                )

            with tab_xlsx:
                try:
                    if st.button("Gerar XLSX"):
                        buffer_xlsx = _gerar_xlsx(st.session_state.calendario_gerado, st.session_state.mes_ano_calendario)
                        st.download_button(
                            "Baixar XLSX",
                            data=buffer_xlsx.getvalue(),
                            file_name=f"calendario_{st.session_state.mes_ano_calendario.replace(' ', '_').lower()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                except ImportError:
                    st.write("Para XLSX: pip install openpyxl")
                except Exception as e:
                    st.error(f"Erro XLSX: {str(e)}")


def _gerar_xlsx(calendario_csv: str, mes_ano: str):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Calendário {mes_ano}"

    ws.merge_cells('A1:G1')
    ws['A1'] = f"CALENDÁRIO - {mes_ano}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    dias_semana = ["DOMINGO", "SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SÁBADO"]
    for col, dia in enumerate(dias_semana, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = dia
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    linha_atual = 4
    for linha in calendario_csv.split('\n'):
        if linha.strip() and not linha.startswith(',,'):
            celulas = linha.split(',')
            for col, conteudo in enumerate(celulas, 1):
                if conteudo.strip():
                    cell = ws.cell(row=linha_atual, column=col)
                    cell.value = conteudo.strip()
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    thin = Side(style='thin')
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            linha_atual += 1

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 30
        for row in range(4, linha_atual):
            ws.row_dimensions[row].height = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
