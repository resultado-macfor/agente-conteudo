"""
Página de Criação de Calendário Editorial.
Geração de calendário de pautas para conteúdo agrícola.
"""
import streamlit as st
import datetime
from io import BytesIO
from database import modelo_texto
from models import construir_contexto
from fallback import get_dados_safra_completos


def get_conhecimento_safras(contexto_agente):
    """
    Retorna o conhecimento sobre safras agrícolas.
    Prioriza dados do agente, usa fallback hardcoded se não existir.

    Args:
        contexto_agente: Contexto construído do agente

    Returns:
        String com dados de safra (do agente ou fallback)
    """
    indicadores_safra = ['DADOS_SAFRA', 'Calendário de Safra', 'Plantio de', 'Colheita de']

    tem_dados_safra = any(indicador in contexto_agente for indicador in indicadores_safra)

    if tem_dados_safra:
        return ""  
    else:
        # Usar dados hardcoded como fallback
        return get_dados_safra_completos()


def render():
    """Renderiza a aba de criação de calendário."""
    st.header("📅 Criadora de Calendário")

    if not st.session_state.get('agente_selecionado'):
        st.warning("Nenhum agente selecionado.")
        return

    agente = st.session_state.agente_selecionado
    st.success(f"Agente: {agente['nome']}")

    col1, col2 = st.columns([2, 1])

    # Lista de estados brasileiros
    ESTADOS_BRASIL = [
        "Todos os estados",
        "Acre", "Alagoas", "Amapá", "Amazonas", "Bahia", "Ceará",
        "Distrito Federal", "Espírito Santo", "Goiás", "Maranhão",
        "Mato Grosso", "Mato Grosso do Sul", "Minas Gerais", "Pará",
        "Paraíba", "Paraná", "Pernambuco", "Piauí", "Rio de Janeiro",
        "Rio Grande do Norte", "Rio Grande do Sul", "Rondônia", "Roraima",
        "Santa Catarina", "São Paulo", "Sergipe", "Tocantins"
    ]

    with col1:
        mes_ano = st.text_input("Mês/Ano:", "FEVEREIRO 2026")
        data_inicio = st.date_input("Data início:", value=datetime.date(2026, 2, 1))
        data_fim = st.date_input("Data fim:", value=datetime.date(2026, 2, 28))

        delta_dias = (data_fim - data_inicio).days + 1

        # Seleção de estado
        estado_selecionado = st.selectbox(
            "Estado de referência para safras:",
            ESTADOS_BRASIL,
            index=0,
            help="Selecione o estado para considerar o calendário de safras específico da região"
        )

        culturas_prioritarias = st.text_area(
            "Culturas (separadas por vírgula):",
            "Soja, Milho, Cana-de-açúcar, Algodão"
        )
        culturas_lista = [c.strip() for c in culturas_prioritarias.split(",") if c.strip()]

    with col2:
        dias_com_1_pauta = st.number_input("Dias com 1 pauta:", 0, delta_dias, 5)
        dias_com_2_pautas = st.number_input("Dias com 2 pautas:", 0, delta_dias, 15)
        dias_com_3_pautas = st.number_input("Dias com 3 pautas:", 0, delta_dias, 3)
        dias_sem_pautas = delta_dias - (dias_com_1_pauta + dias_com_2_pautas + dias_com_3_pautas)

        if dias_sem_pautas < 0:
            st.error("Total excede dias disponíveis")

    st.subheader("Produtos e Direcionais")
    st.write("Formato: Produto(s) - Cultura(s) - Tema")

    produtos_direcionais = st.text_area(
        "Produtos com culturas e temas:",
        """Verdavis, Megafol e Victrato - Soja e Milho - Tecnologia para feira
        Elestal Neo - Soja - Controle de mosca-branca
        Fortenza - Milho - Seedcare para cigarrinha
        YieldOn - Soja - Bioativador para pegamento
        Miravis - Soja - Fungicida para ferrugem""",
        height=150
    )

    produtos_com_direcionais = []
    if produtos_direcionais:
        for linha in produtos_direcionais.split('\n'):
            linha = linha.strip()
            if linha and ' - ' in linha:
                partes = linha.split(' - ')
                if len(partes) >= 3:
                    produtos = [p.strip() for p in partes[0].split(' e ') if p.strip()]
                    if len(produtos) == 1 and ',' in produtos[0]:
                        produtos = [p.strip() for p in produtos[0].split(',') if p.strip()]
                    culturas = [c.strip() for c in partes[1].split(' e ') if c.strip()]
                    tema = ' - '.join(partes[2:]).strip()
                    produtos_com_direcionais.append({
                        'produtos': produtos,
                        'culturas': culturas,
                        'tema': tema
                    })

    col_feira, col_recorrente = st.columns(2)

    with col_feira:
        st.write("Semana com evento (1 post/dia):")
        semana_feira_inicio = st.date_input("Início evento:", value=datetime.date(2026, 2, 9))
        semana_feira_fim = st.date_input("Fim evento:", value=datetime.date(2026, 2, 13))
        produtos_prioritarios_feira = st.text_input("Produtos prioritários:", "Verdavis, Megafol, Victrato")

    with col_recorrente:
        pauta_recorrente_texto = st.text_input("Pauta fixa:", "Victrato pelo Brasil")
        pauta_recorrente_dias = st.multiselect(
            "Dias da semana:",
            ["Terça", "Quinta"],
            default=["Terça", "Quinta"]
        )

    contexto_mensal = st.text_area(
        "Contexto do mês:",
        """FEVEREIRO 2026:
- Soja: colheita no centro-sul
- Milho: plantio da safrinha
- Cana: crescimento vegetativo
- Evento: Feira Nacional do Agronegócio (09-13/02)""",
        height=120
    )

    with st.expander("⚙️ Configurações Avançadas"):
        evitar_consecutivos = st.checkbox("Evitar dias consecutivos sem pautas", True)
        max_repeticoes = st.slider("Máx repetições por tema:", 1, 5, 2)
        biologicos_fim_semana = st.checkbox("Priorizar biológicos nos fins de semana", False)

    if st.button("Gerar Calendário", type="primary"):
        if data_inicio >= data_fim:
            st.error("Data início deve ser anterior")
        elif not culturas_lista:
            st.error("Digite culturas")
        elif (dias_com_1_pauta + dias_com_2_pautas + dias_com_3_pautas) > delta_dias:
            st.error("Total excede período")
        else:
            with st.spinner("Gerando calendário..."):
                try:
                    contexto_agente = construir_contexto(agente, st.session_state.get('segmentos_selecionados', []))

                    info_especifica = f"""
CONFIGURAÇÕES DO CALENDÁRIO:

1. SEMANA COM EVENTO ({semana_feira_inicio.strftime('%d/%m')} a {semana_feira_fim.strftime('%d/%m')}):
   - APENAS 1 pauta por dia nesta semana
   - Priorizar produtos: {produtos_prioritarios_feira}

2. PAUTA FIXA OBRIGATÓRIA: "{pauta_recorrente_texto}"
   - Dias: {', '.join(pauta_recorrente_dias)}
   - Esta pauta DEVE aparecer em TODOS esses dias

3. DISTRIBUIÇÃO DE PAUTAS:
   - Dias com 1 pauta: {dias_com_1_pauta}
   - Dias com 2 pautas: {dias_com_2_pautas}
   - Dias com 3 pautas: {dias_com_3_pautas}
   - Evitar consecutivos sem pauta: {evitar_consecutivos}

4. CONTROLE DE REPETIÇÃO:
   - Máximo repetições por tema: {max_repeticoes}
   - Células podem ter múltiplas culturas/produtos
"""
                    if biologicos_fim_semana:
                        info_especifica += """
5. BIOLÓGICOS NOS FINS DE SEMANA:
   - Sábados e Domingos devem priorizar produtos biológicos
"""
                    # Conhecimento de safras 
                    conhecimento_safras = get_conhecimento_safras(contexto_agente)
                    prompt_calendario = f'''
{contexto_agente}

{conhecimento_safras}

GERAR CALENDÁRIO COM ESTAS REGRAS:
PERÍODO: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}
MÊS: {mes_ano}
ESTADO DE REFERÊNCIA: {estado_selecionado}
{f"IMPORTANTE: Use os dados de safra específicos para {estado_selecionado}. Considere as datas de plantio e colheita deste estado." if estado_selecionado != "Todos os estados" else "Use os dados de safra considerando todas as regiões do Brasil."}

{info_especifica}

CONTEXTO DO MÊS: {contexto_mensal}

PRODUTOS E TEMAS FORNECIDOS (USE EXATAMENTE COMO ESCRITOS):
{chr(10).join([f"- {', '.join(p['produtos'])} - {', '.join(p['culturas'])} - {p['tema']}" for p in produtos_com_direcionais])}

=== REGRAS CRÍTICAS (SIGA RIGOROSAMENTE) ===

1. Semana {semana_feira_inicio.strftime('%d/%m')} a {semana_feira_fim.strftime('%d/%m')}: APENAS 1 PAUTA POR DIA
2. Priorizar produtos: {produtos_prioritarios_feira} na semana do evento
3. Inserir "{pauta_recorrente_texto}" em TODAS as {', '.join(pauta_recorrente_dias)} - SEM EXCEÇÃO
4. NÃO repetir o mesmo tema mais de {max_repeticoes} vezes no mês
5. Células PODEM e DEVEM ter múltiplas culturas juntas: "Soja e Milho" é UM tema único
6. Praticamente todos os dias devem ter conteúdo
7. NUNCA deixar 3 dias consecutivos sem pautas
8. Baseie as pautas no contexto do mês fornecido
9. Respeitar COM RIGIDEZ as fases reais de cada cultura por estado (ver DADOS_SAFRA)

=== REGRAS DE FIDELIDADE AOS TEMAS ===

10. NÃO MODIFIQUE os temas fornecidos - use EXATAMENTE como foram escritos
11. NÃO CRIE temas ou produtos que não foram solicitados
12. "Soja e Milho" ou "Soja/Milho" é UMA pauta única - NUNCA divida em duas linhas separadas
13. O tema já contém a descrição completa - NÃO adicione palavras extras
14. Use apenas os produtos listados em "PRODUTOS E TEMAS FORNECIDOS"

=== FORMATO DE SAÍDA (OBRIGATÓRIO) ===

RETORNE APENAS O CSV PURO. NÃO inclua:
- Explicações ou justificativas
- Observações sobre as regras
- Comentários sobre conflitos
- Textos introdutórios ou conclusivos
- Markdown (###, **, etc)

Formato da célula: "[EMOJI] Produto(s) - Cultura(s) - Tema"

EXEMPLO DE SAÍDA CORRETA (apenas isso, nada mais):
DOMINGO,SEGUNDA,TERÇA,QUARTA,QUINTA,SEXTA,SÁBADO
,🔵 Verdavis - Soja - Tema,🟢 Victrato pelo Brasil,🔵 Produto - Cultura - Tema,🟢 Victrato pelo Brasil,🔵 Produto - Cultura - Tema,
🟡 Produto - Cultura - Tema,🔵 Produto - Cultura - Tema,🟢 Victrato pelo Brasil,...

Retorne SOMENTE o CSV com 7 colunas (Domingo a Sábado), sem nenhum texto adicional.
'''

                    resposta = modelo_texto.generate_content(prompt_calendario)
                    calendario_csv = resposta.text

                    calendario_limpo = calendario_csv.strip()
                    if '```csv' in calendario_limpo:
                        calendario_limpo = calendario_limpo.replace('```csv', '').replace('```', '')
                    if '```' in calendario_limpo:
                        calendario_limpo = calendario_limpo.replace('```', '')

                    st.session_state.calendario_gerado = calendario_limpo
                    st.session_state.mes_ano_calendario = mes_ano

                    st.success("Calendário gerado")

                except Exception as e:
                    st.error(f"Erro: {str(e)}")

    if 'calendario_gerado' in st.session_state:
        st.subheader(f"Calendário - {st.session_state.mes_ano_calendario}")

        tab_csv, tab_xlsx = st.tabs(["CSV", "XLSX"])

        with tab_csv:
            st.text_area("CSV:", st.session_state.calendario_gerado, height=400)

            st.download_button(
                "Baixar CSV",
                data=st.session_state.calendario_gerado,
                file_name=f"calendario_{mes_ano.replace(' ', '_').lower()}.csv",
                mime="text/csv"
            )

        with tab_xlsx:
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side

                def gerar_xlsx():
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = f"Calendário {mes_ano}"

                    ws.merge_cells('A1:G1')
                    ws['A1'] = f"CALENDÁRIO - {mes_ano}"
                    ws['A1'].font = Font(bold=True, size=14)
                    ws['A1'].alignment = Alignment(horizontal='center')

                    dias_semana = ["DOMINGO", "SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SÁBADO"]
                    for col, dia in enumerate(dias_semana, 1):
                        cell = ws.cell(row=3, column=col)
                        cell.value = dia
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                    linhas = st.session_state.calendario_gerado.split('\n')
                    linha_atual = 4

                    for linha in linhas:
                        if linha.strip() and not linha.startswith(',,'):
                            celulas = linha.split(',')
                            for col, conteudo in enumerate(celulas, 1):
                                if conteudo.strip():
                                    cell = ws.cell(row=linha_atual, column=col)
                                    cell.value = conteudo.strip()
                                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    cell.border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                            linha_atual += 1

                    for col in range(1, 8):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
                        for row in range(4, linha_atual):
                            ws.row_dimensions[row].height = 60

                    buffer = BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    return buffer

                if st.button("Gerar XLSX"):
                    buffer_xlsx = gerar_xlsx()

                    st.download_button(
                        "Baixar XLSX",
                        data=buffer_xlsx.getvalue(),
                        file_name=f"calendario_{mes_ano.replace(' ', '_').lower()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except ImportError:
                st.write("Para XLSX: pip install openpyxl")
            except Exception as e:
                st.error(f"Erro XLSX: {str(e)}")
